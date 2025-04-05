from io import BytesIO
from typing import Sequence
from dataclasses import fields

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from src.expense_manager.types import Expense


def expenses_to_xlsx_file(expenses: Sequence[Expense]) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    column_titles = [field.name for field in fields(Expense)]  # noqa

    ws.append(column_titles)

    for expense in expenses:
        ws.append((
            expense.id,
            expense.name,
            float(expense.amount_uah),
            float(expense.amount_usd),
            expense.created_at.split("T")[0],
        ))

    ws.append((
        "",
        "Загалом",
        f"=SUM(C2:C{len(expenses)})",
        f"=SUM(D2:D{len(expenses)})"
    ))

    for cell in ws[1] + ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
